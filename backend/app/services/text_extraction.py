from __future__ import annotations

import io
from pathlib import Path
from typing import Any

from app.core.encryption import decrypt
from app.utils.errors import raise_error

MAX_TEXT_LENGTH = 60_000
XLSX_MAX_SHEETS = 2
XLSX_MAX_ROWS = 200
XLSX_TABLE_MAX_ROWS = 20
DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
PDF_MIME = "application/pdf"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _read_decrypted(path: str) -> bytes:
    p = Path(path)
    if not p.exists():
        raise_error(400, "BAD_REQUEST", "File not found", {"path": path})
    return decrypt(p.read_bytes())


def extract_text_from_pdf(path: str) -> str:
    return extract_pdf_payload(path)["full_text"]


def extract_pdf_payload(path: str) -> dict[str, Any]:
    try:
        from pypdf import PdfReader
    except ImportError:
        raise_error(500, "INTERNAL_ERROR", "PDF support not available", {})
    data = _read_decrypted(path)
    try:
        reader = PdfReader(io.BytesIO(data))
        parts: list[str] = []
        page_breaks: list[int] = []
        total = 0
        for page_idx, page in enumerate(reader.pages):
            if total >= MAX_TEXT_LENGTH:
                break
            if page_idx > 0:
                page_breaks.append(total)
            text = page.extract_text() or ""
            remaining = MAX_TEXT_LENGTH - total
            if len(text) > remaining:
                text = text[:remaining]
            parts.append(text)
            total += len(text)
        raw = "".join(parts).strip()
        if not raw:
            raise_error(400, "BAD_REQUEST", "Cannot read text from document. The file may be empty or image-only.", {})
        return {
            "full_text": raw,
            "rich_content": None,
            "source_format": "pdf",
            "page_breaks": [offset for offset in page_breaks if 0 < offset <= len(raw)],
        }
    except Exception as e:
        raise_error(400, "BAD_REQUEST", "Cannot read text from document.", {"detail": str(e)})


def extract_text_from_docx(path: str) -> str:
    rich_payload = extract_docx_rich_payload(path)
    return rich_payload["full_text"]


def _truncate_text(value: str) -> str:
    return value[:MAX_TEXT_LENGTH] if len(value) > MAX_TEXT_LENGTH else value


def _docx_alignment_to_text(value: Any) -> str | None:
    if value is None:
        return None
    mapping = {
        "LEFT": "left",
        "CENTER": "center",
        "RIGHT": "right",
        "JUSTIFY": "justify",
    }
    return mapping.get(getattr(value, "name", ""), None)


def _iter_docx_blocks(document: Any):
    try:
        from docx.document import Document as DocxDocumentType
        from docx.table import Table, _Cell
        from docx.text.paragraph import Paragraph
        from docx.oxml.table import CT_Tbl
        from docx.oxml.text.paragraph import CT_P
    except ImportError:
        return []

    parent_elm = document.element.body if isinstance(document, DocxDocumentType) else document._tc
    for child in parent_elm.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, document)
        elif isinstance(child, CT_Tbl):
            yield Table(child, document)


def _text_marks_from_run(run: Any) -> list[dict[str, Any]]:
    marks: list[dict[str, Any]] = []
    if run.bold:
        marks.append({"type": "bold"})
    if run.italic:
        marks.append({"type": "italic"})
    if run.underline:
        marks.append({"type": "underline"})
    if run.font.strike:
        marks.append({"type": "strike"})

    style_attrs: dict[str, str] = {}
    if run.font.name:
        style_attrs["fontFamily"] = run.font.name
    if run.font.size is not None:
        try:
            style_attrs["fontSize"] = f"{int(round(run.font.size.pt))}px"
        except Exception:
            pass
    if style_attrs:
        marks.append({"type": "textStyle", "attrs": style_attrs})
    return marks


def _text_node_from_run(run: Any) -> dict[str, Any] | None:
    text = str(run.text or "").replace("\r", "")
    if not text:
        return None
    node: dict[str, Any] = {"type": "text", "text": text}
    marks = _text_marks_from_run(run)
    if marks:
        node["marks"] = marks
    return node


def _is_list_style(style_name: str) -> tuple[str, bool]:
    lower = style_name.lower()
    if "list bullet" in lower:
        return "bulletList", True
    if "list number" in lower:
        return "orderedList", True
    return "paragraph", False


_W_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"


def _paragraph_has_page_break(paragraph: Any) -> bool:
    try:
        element = paragraph._p
    except AttributeError:
        return False
    for br in element.iter(f"{_W_NS}br"):
        if br.get(f"{_W_NS}type") == "page":
            return True
    if next(element.iter(f"{_W_NS}lastRenderedPageBreak"), None) is not None:
        return True
    return False


def _paragraph_node_from_docx(paragraph: Any) -> tuple[dict[str, Any], str | None]:
    style_name = str(getattr(getattr(paragraph, "style", None), "name", "") or "")
    heading_level = None
    if style_name.lower().startswith("heading "):
        try:
            heading_level = int(style_name.split()[-1])
        except Exception:
            heading_level = 1

    node_type = "heading" if heading_level else "paragraph"
    attrs: dict[str, Any] = {}
    if heading_level:
        attrs["level"] = max(1, min(3, heading_level))

    alignment = _docx_alignment_to_text(paragraph.alignment)
    if alignment:
        attrs["textAlign"] = alignment

    content = [
        text_node
        for run in paragraph.runs
        if (text_node := _text_node_from_run(run)) is not None
    ]
    if not content:
        content = [{"type": "text", "text": paragraph.text or ""}] if paragraph.text else []

    node: dict[str, Any] = {"type": node_type}
    if attrs:
        node["attrs"] = attrs
    if content:
        node["content"] = content

    list_type, is_list = _is_list_style(style_name)
    return node, list_type if is_list else None


def _table_node_from_docx(table: Any) -> dict[str, Any] | None:
    rows_data = _extract_docx_table_rows(table)
    if not rows_data:
        return None

    if _looks_like_layout_table(rows_data):
        return {
            "type": "doc",
            "content": _layout_table_to_paragraphs(rows_data),
        }

    rows: list[dict[str, Any]] = []
    for row_idx, row_cells_data in enumerate(rows_data):
        row_cells: list[dict[str, Any]] = []
        for cell_content in row_cells_data:
            row_cells.append(
                {
                    "type": "tableHeader" if row_idx == 0 else "tableCell",
                    "content": cell_content,
                }
            )
        rows.append({"type": "tableRow", "content": row_cells})
    if not rows:
        return None
    return {"type": "table", "content": rows}


def _extract_docx_table_rows(table: Any) -> list[list[list[dict[str, Any]]]]:
    rows_data: list[list[list[dict[str, Any]]]] = []
    for row in table.rows:
        row_cells_data: list[list[dict[str, Any]]] = []
        seen_cells: set[int] = set()
        for cell in row.cells:
            cell_key = id(getattr(cell, "_tc", cell))
            if cell_key in seen_cells:
                continue
            seen_cells.add(cell_key)

            cell_content: list[dict[str, Any]] = []
            for block in _iter_docx_blocks(cell):
                if hasattr(block, "runs"):
                    paragraph_node, _ = _paragraph_node_from_docx(block)
                    if paragraph_node.get("content"):
                        cell_content.append(paragraph_node)
            if not cell_content:
                cell_content = [{"type": "paragraph"}]
            row_cells_data.append(cell_content)
        if row_cells_data:
            rows_data.append(row_cells_data)
    return rows_data


def _cell_text_from_content(cell_content: list[dict[str, Any]]) -> str:
    return " ".join(
        _serialize_rich_node_to_text(node).strip()
        for node in cell_content
        if _serialize_rich_node_to_text(node).strip()
    ).strip()


def _looks_like_layout_table(rows_data: list[list[list[dict[str, Any]]]]) -> bool:
    flat_cells = [cell for row in rows_data for cell in row]
    if not flat_cells:
        return False

    text_cells = [_cell_text_from_content(cell) for cell in flat_cells]
    non_empty = [value for value in text_cells if value]
    if not non_empty:
        return True

    max_cols = max((len(row) for row in rows_data), default=0)
    empty_ratio = 1 - (len(non_empty) / max(1, len(flat_cells)))
    unique_ratio = len(set(non_empty)) / max(1, len(non_empty))

    return max_cols >= 6 or empty_ratio >= 0.45 or unique_ratio <= 0.55


def _layout_table_to_paragraphs(rows_data: list[list[list[dict[str, Any]]]]) -> list[dict[str, Any]]:
    paragraphs: list[dict[str, Any]] = []
    for row in rows_data:
        parts: list[str] = []
        for cell in row:
            text = _cell_text_from_content(cell)
            if text and (not parts or parts[-1] != text):
                parts.append(text)
        line = " ".join(parts).strip()
        if line:
            paragraphs.append(
                {
                    "type": "paragraph",
                    "content": [{"type": "text", "text": line}],
                }
            )
    return paragraphs or [{"type": "paragraph"}]


def _serialize_rich_node_to_text(node: dict[str, Any] | None) -> str:
    if not isinstance(node, dict):
        return ""

    node_type = node.get("type")
    content = node.get("content")

    if node_type == "text":
        return str(node.get("text") or "")

    if not isinstance(content, list):
        return ""

    separator = ""
    if node_type in {"doc", "bulletList", "orderedList", "listItem", "table"}:
        separator = "\n"
    elif node_type == "tableRow":
        separator = " | "
    elif node_type in {"tableCell", "tableHeader"}:
        separator = "\n"

    parts = [_serialize_rich_node_to_text(child) for child in content]
    return separator.join(part for part in parts if part)


def _compute_doc_page_breaks(
    rich_content: dict[str, Any], block_breaks_before: list[bool]
) -> list[int]:
    children = rich_content.get("content") or []
    if len(block_breaks_before) != len(children):
        return []

    breaks: list[int] = []
    serialized_parts: list[str] = []
    for child, has_break in zip(children, block_breaks_before):
        text = _serialize_rich_node_to_text(child)
        if not text:
            continue
        if serialized_parts and has_break:
            current = "\n".join(serialized_parts)
            breaks.append(len(current) + 1)
        serialized_parts.append(text)
    return breaks


def extract_docx_rich_payload(path: str) -> dict[str, Any]:
    try:
        from docx import Document as DocxDocument
        from docx.table import Table
    except ImportError:
        raise_error(500, "INTERNAL_ERROR", "DOCX support not available", {})

    data = _read_decrypted(path)

    try:
        doc = DocxDocument(io.BytesIO(data))
        content: list[dict[str, Any]] = []
        block_breaks_before: list[bool] = []
        pending_list_type: str | None = None
        pending_list_items: list[dict[str, Any]] = []
        pending_break_before_list = False
        pending_break_in_list = False

        def flush_pending_list() -> None:
            nonlocal pending_list_type, pending_list_items, pending_break_before_list, pending_break_in_list
            if pending_list_type and pending_list_items:
                content.append({"type": pending_list_type, "content": pending_list_items})
                block_breaks_before.append(pending_break_before_list)
            pending_list_type = None
            pending_list_items = []
            pending_break_before_list = pending_break_in_list
            pending_break_in_list = False

        for block in _iter_docx_blocks(doc):
            if hasattr(block, "rows") and isinstance(block, Table):
                flush_pending_list()
                table_node = _table_node_from_docx(block)
                if table_node is not None:
                    if table_node.get("type") == "doc" and isinstance(table_node.get("content"), list):
                        for child in table_node["content"]:
                            content.append(child)
                            block_breaks_before.append(False)
                    else:
                        content.append(table_node)
                        block_breaks_before.append(False)
                continue

            has_break = _paragraph_has_page_break(block)
            paragraph_node, list_type = _paragraph_node_from_docx(block)
            if list_type:
                if pending_list_type != list_type:
                    flush_pending_list()
                    pending_list_type = list_type
                if has_break:
                    if not pending_list_items:
                        pending_break_before_list = True
                    else:
                        pending_break_in_list = True
                pending_list_items.append({"type": "listItem", "content": [paragraph_node]})
            else:
                flush_pending_list()
                content.append(paragraph_node)
                block_breaks_before.append(has_break)

        flush_pending_list()

        rich_content = {"type": "doc", "content": content or [{"type": "paragraph"}]}
        page_breaks = _compute_doc_page_breaks(rich_content, block_breaks_before)
        raw_text = _serialize_rich_node_to_text(rich_content)
        stripped_text = raw_text.strip()
        leading_strip = len(raw_text) - len(raw_text.lstrip())
        if leading_strip:
            page_breaks = [pb - leading_strip for pb in page_breaks if pb > leading_strip]
        full_text = _truncate_text(stripped_text)
        page_breaks = [pb for pb in page_breaks if 0 < pb < len(full_text)]
        if not full_text:
            raise_error(400, "BAD_REQUEST", "Cannot read text from document. The file may be empty.", {})

        return {
            "full_text": full_text,
            "rich_content": rich_content,
            "source_format": "docx",
            "page_breaks": page_breaks,
        }
    except Exception as e:
        raise_error(400, "BAD_REQUEST", "Cannot read text from document.", {"detail": str(e)})


def extract_text_from_xlsx(path: str) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise_error(500, "INTERNAL_ERROR", "XLSX support not available", {})
    data = _read_decrypted(path)
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        parts: list[str] = []
        total = 0
        for idx, sheet in enumerate(wb.worksheets):
            if idx >= XLSX_MAX_SHEETS:
                break
            sheet_name = sheet.title
            row_count = 0
            for row in sheet.iter_rows(values_only=True):
                if row_count >= XLSX_MAX_ROWS or total >= MAX_TEXT_LENGTH:
                    break
                line = " ".join(str(c) if c is not None else "" for c in row).strip()
                if line:
                    parts.append(f"Sheet: {sheet_name} Row: {line}")
                    total += len(parts[-1]) + 1
                row_count += 1
        wb.close()
        raw = "\n".join(parts).strip()
        if not raw:
            raise_error(400, "BAD_REQUEST", "Cannot read text from XLSX. The file may be empty.", {})
        return raw[:MAX_TEXT_LENGTH]
    except Exception as e:
        raise_error(400, "BAD_REQUEST", "Cannot read text from document.", {"detail": str(e)})


def extract_tables_from_xlsx(path: str) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    p = Path(path)
    if not p.exists():
        return []
    try:
        data = decrypt(p.read_bytes())
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        out: list[dict[str, Any]] = []
        for idx, sheet in enumerate(wb.worksheets):
            if idx >= 1:
                break
            rows: list[list[str]] = []
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_idx >= XLSX_TABLE_MAX_ROWS:
                    break
                rows.append([str(c) if c is not None else "" for c in row])
            if rows:
                out.append({"name": sheet.title or "Sheet 1", "rows": rows})
        wb.close()
        return out
    except Exception:
        return []


def extract_text(path: str, mime_type: str) -> str:
    if mime_type == PDF_MIME:
        return extract_text_from_pdf(path)
    if mime_type == DOCX_MIME:
        return extract_text_from_docx(path)
    if mime_type == XLSX_MIME:
        return extract_text_from_xlsx(path)
    raise_error(
        400,
        "BAD_REQUEST",
        "Unsupported file type for text extraction. Use PDF, DOCX or XLSX.",
        {"mime_type": mime_type},
    )


def extract_advanced_editor_payload(path: str, mime_type: str) -> dict[str, Any]:
    if mime_type == DOCX_MIME:
        return extract_docx_rich_payload(path)
    if mime_type == PDF_MIME:
        return extract_pdf_payload(path)
    return {
        "full_text": extract_text(path, mime_type),
        "rich_content": None,
        "source_format": "plain_text",
        "page_breaks": [],
    }
